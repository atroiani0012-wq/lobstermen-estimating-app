"""
Output 5: 05_Bidder_List.xlsx — prospective bidders discovered during web research.

Generated only when `analysis['web_research']['bidder_list']` has entries.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="111827")
SUBTITLE_FONT = Font(italic=True, size=10, color="6B7280")


def build_bidder_list_xlsx(analysis: dict[str, Any]) -> bytes:
    wr = analysis.get("web_research") or {}
    bidders = wr.get("bidder_list") or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Bidders"

    project = analysis.get("project") or {}
    ws["A1"] = f"Prospective Bidders — {project.get('name', 'Unknown Project')}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        f"{project.get('location', '')} · {project.get('client', '')} · "
        f"Source: UMA Estimating app web research"
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    headers = ["Company", "Location", "Contact Name", "Email", "Phone", "Source URL"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")

    row = 5
    for b in bidders:
        ws.cell(row=row, column=1, value=b.get("company_name", ""))
        ws.cell(row=row, column=2, value=b.get("location", ""))
        ws.cell(row=row, column=3, value=b.get("contact_name", ""))
        ws.cell(row=row, column=4, value=b.get("contact_email", ""))
        ws.cell(row=row, column=5, value=b.get("contact_phone", ""))
        ws.cell(row=row, column=6, value=b.get("source_url", ""))
        row += 1

    # Auto-width
    widths = [32, 22, 22, 28, 18, 48]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze top rows
    ws.freeze_panes = "A5"

    # Findings sheet
    findings = wr.get("findings") or []
    if findings:
        ws2 = wb.create_sheet("Research Findings")
        ws2["A1"] = f"Research Findings — {project.get('name', '')}"
        ws2["A1"].font = TITLE_FONT
        ws2.merge_cells("A1:C1")

        f_headers = ["Topic", "Summary", "Source URL"]
        for col_idx, h in enumerate(f_headers, start=1):
            c = ws2.cell(row=3, column=col_idx, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT

        row = 4
        for f in findings:
            ws2.cell(row=row, column=1, value=f.get("topic", ""))
            ws2.cell(row=row, column=2, value=f.get("summary", ""))
            ws2.cell(row=row, column=3, value=f.get("source_url", ""))
            ws2.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

        ws2.column_dimensions["A"].width = 24
        ws2.column_dimensions["B"].width = 80
        ws2.column_dimensions["C"].width = 48
        ws2.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
