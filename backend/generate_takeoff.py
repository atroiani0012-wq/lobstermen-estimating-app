"""
Output 1 — Takeoff/Quantity spreadsheet in UMA template layout.

This is NOT the full UMA master estimator (which has cost formulas, labor/per-diem/etc.
baked in). It's a **takeoff-only** sheet the estimator can paste into the master.

Layout mirrors the project-header + item-table pattern used across UMA master templates
(HDPR, Shotcrete, Soil Nail, Pre-drilling, Micropile).
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", start_color="1E3A8A")
_SUBHEADER_FILL = PatternFill("solid", start_color="D5E8F0")
_THIN = Side(style="thin", color="888888")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Source × confidence → fill color.
#   green  = user_measured OR dimension_read at high confidence (trustworthy)
#   yellow = vision_detected at medium+ confidence (plausible, spot-check)
#   red    = estimated OR low confidence (needs review)
_FILL_GREEN  = PatternFill("solid", start_color="D1FAE5")  # soft green
_FILL_YELLOW = PatternFill("solid", start_color="FEF3C7")  # soft amber
_FILL_RED    = PatternFill("solid", start_color="FEE2E2")  # soft red
_FILL_GREY   = PatternFill("solid", start_color="F3F4F6")  # fallback

_TRUSTED_SOURCES = {"user_measured", "dimension_read"}


def _confidence_fill(source: Any, confidence: Any) -> PatternFill:
    src = (str(source) if source is not None else "").lower()
    conf = (str(confidence) if confidence is not None else "").lower()
    if conf == "low" or src == "estimated":
        return _FILL_RED
    if src in _TRUSTED_SOURCES and conf == "high":
        return _FILL_GREEN
    if src == "vision_detected" and conf in ("high", "medium"):
        return _FILL_YELLOW
    if not src and not conf:
        return _FILL_GREY
    return _FILL_YELLOW  # any other mid-case: caution tint


def _write_header_block(ws, project: dict[str, Any]) -> int:
    rows = [
        ("Project Name", project.get("name", "")),
        ("Location", project.get("location", "")),
        ("Client / GC", project.get("client", "")),
        ("Bid Due Date", project.get("bid_due_date", "")),
        ("Scope Type", project.get("scope_type", "")),
        ("Description", project.get("description", "")),
        ("Site Conditions", project.get("site_conditions", "")),
        ("Schedule Notes", project.get("schedule_notes", "")),
    ]
    for r, (label, value) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, color="FFFFFF")
        ws.cell(row=r, column=1).fill = _HEADER_FILL
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
        ws.cell(row=r, column=2, value=str(value)).font = Font(name="Arial")
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    return len(rows) + 2  # first blank row after header


def _write_items_table(ws, start_row: int, title: str, items: list[dict[str, Any]], cols: list[tuple[str, str, int]]) -> int:
    ws.cell(row=start_row, column=1, value=title).font = Font(name="Arial", bold=True, size=12)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(cols))
    start_row += 1
    for c, (label, _key, width) in enumerate(cols, start=1):
        cell = ws.cell(row=start_row, column=c, value=label)
        cell.font = Font(name="Arial", bold=True)
        cell.fill = _SUBHEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = width
    header_row = start_row
    start_row += 1
    for item in items or []:
        row_fill = _confidence_fill(item.get("source"), item.get("confidence"))
        for c, (_label, key, _w) in enumerate(cols, start=1):
            val = item.get(key, "")
            if key in ("quantity", "unit_cost_est") and isinstance(val, (int, float)):
                cell = ws.cell(row=start_row, column=c, value=val)
                cell.number_format = '#,##0.00' if key == "unit_cost_est" else '#,##0.##'
            else:
                display = val
                if isinstance(display, (int, float)) and display == 0:
                    display = ""
                cell = ws.cell(row=start_row, column=c, value=display)
            cell.font = Font(name="Arial")
            cell.border = _BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Highlight the source + confidence cells (and optionally the whole row
            # tint for at-a-glance scanning) so Tony can spot red rows instantly.
            if key in ("source", "confidence"):
                cell.fill = row_fill
        start_row += 1
    # Add a total row with formula if we have qty & unit cost columns
    qty_col = next((i for i, (_, k, _) in enumerate(cols, start=1) if k == "quantity"), None)
    uc_col = next((i for i, (_, k, _) in enumerate(cols, start=1) if k == "unit_cost_est"), None)
    if qty_col and uc_col and items:
        total_col = len(cols) + 1
        ws.cell(row=header_row, column=total_col, value="Extended").font = Font(name="Arial", bold=True)
        ws.cell(row=header_row, column=total_col).fill = _SUBHEADER_FILL
        ws.cell(row=header_row, column=total_col).border = _BORDER
        ws.column_dimensions[get_column_letter(total_col)].width = 14
        for r in range(header_row + 1, start_row):
            q = f"{get_column_letter(qty_col)}{r}"
            u = f"{get_column_letter(uc_col)}{r}"
            cell = ws.cell(row=r, column=total_col, value=f"=IFERROR({q}*{u},0)")
            cell.number_format = '"$"#,##0.00'
            cell.border = _BORDER
            cell.font = Font(name="Arial")
        total_row = start_row
        ws.cell(row=total_row, column=total_col - 1, value="TOTAL").font = Font(name="Arial", bold=True)
        ws.cell(row=total_row, column=total_col - 1).alignment = Alignment(horizontal="right")
        sum_range = f"{get_column_letter(total_col)}{header_row + 1}:{get_column_letter(total_col)}{start_row - 1}"
        cell = ws.cell(row=total_row, column=total_col, value=f"=SUM({sum_range})")
        cell.font = Font(name="Arial", bold=True)
        cell.number_format = '"$"#,##0.00'
        cell.border = _BORDER
        cell.fill = _SUBHEADER_FILL
        start_row += 1
    return start_row + 2


def build_takeoff_xlsx(analysis: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Takeoff"

    project = analysis.get("project") or {}
    next_row = _write_header_block(ws, project)

    next_row = _write_items_table(
        ws, next_row, "ITEMS / MATERIALS (Production scope)",
        analysis.get("takeoff_items") or [],
        [("Item / Material", "item", 48), ("Unit", "unit", 8),
         ("Quantity", "quantity", 12), ("Unit Cost (est.)", "unit_cost_est", 16),
         ("Notes / Source", "notes", 44),
         ("Sheet", "source_sheet", 10),
         ("Source", "source", 14),
         ("Confidence", "confidence", 12)],
    )

    next_row = _write_items_table(
        ws, next_row, "TESTING REQUIREMENTS (Priced separately)",
        analysis.get("testing_requirements") or [],
        [("Test Type", "test_type", 36), ("Unit", "unit", 8),
         ("Quantity", "quantity", 10), ("Spec Ref", "reference_spec_section", 18),
         ("Notes", "notes", 40),
         ("Source", "source", 14),
         ("Confidence", "confidence", 12)],
    )

    # Equipment summary sheet
    # Legend for the color coding — kept compact so the Takeoff sheet
    # stays scannable. The three cells are colored to match the values.
    legend_row = next_row
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(name="Arial", bold=True)
    cell = ws.cell(row=legend_row, column=2, value="GREEN = trusted (measured / dimension line)")
    cell.fill = _FILL_GREEN; cell.font = Font(name="Arial")
    ws.merge_cells(start_row=legend_row, start_column=2, end_row=legend_row, end_column=4)
    cell = ws.cell(row=legend_row, column=5, value="YELLOW = AI-detected (spot-check)")
    cell.fill = _FILL_YELLOW; cell.font = Font(name="Arial")
    ws.merge_cells(start_row=legend_row, start_column=5, end_row=legend_row, end_column=6)
    cell = ws.cell(row=legend_row, column=7, value="RED = estimated / low confidence (needs review)")
    cell.fill = _FILL_RED; cell.font = Font(name="Arial")
    ws.merge_cells(start_row=legend_row, start_column=7, end_row=legend_row, end_column=8)

    ws2 = wb.create_sheet("Equipment")
    ws2.append(["Equipment", "Duration (days)", "Notes"])
    for c in range(1, 4):
        ws2.cell(row=1, column=c).font = Font(name="Arial", bold=True)
        ws2.cell(row=1, column=c).fill = _SUBHEADER_FILL
    for eq in analysis.get("equipment_list") or []:
        ws2.append([eq.get("equipment", ""), eq.get("duration_days", 0), eq.get("notes", "")])
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 50

    # File appendix sheet
    ws3 = wb.create_sheet("Files")
    ws3.append(["Filename", "Type", "Summary", "Key info extracted"])
    for c in range(1, 5):
        ws3.cell(row=1, column=c).font = Font(name="Arial", bold=True)
        ws3.cell(row=1, column=c).fill = _SUBHEADER_FILL
    for f in analysis.get("file_appendix") or []:
        ws3.append([f.get("filename", ""), f.get("type", ""), f.get("summary", ""), f.get("key_info", "")])
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 48
    ws3.column_dimensions["D"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
